#Import workbook, date, and web scraping packages
from openpyxl import Workbook, load_workbook
from datetime import datetime
from bs4 import BeautifulSoup
import requests
import time
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np

#Function to update the excel sheet
def update_sheet():

    #Open the workbook and load the google sheet
    wb = load_workbook('stock_prices.xlsx')
    ws = wb['bitcoin']

    #Get the current date and time
    now = datetime.now()

    #Append the time and price into the excel workbook
    time_price = [now, get_price()]
    ws.append(time_price)

    #Save the workbook
    wb.save('stock_prices.xlsx')

#Function to pull data from the internet
def get_price ():

    #Pull in the price
    html_text = requests.get('https://coinmarketcap.com/currencies/bitcoin/').text
    soup = BeautifulSoup(html_text, 'lxml')
    price = soup.find('div', class_ = 'sc-55e6b79c-0 bQrYrL price').text
    index_one = (price.find('$')) + 1
    index_two = (price.find('.')) + 3
    price = price[index_one:index_two]

    #Return the price
    return price

#Function to raph the price in matplotlib
def graph_price ():
    wb = load_workbook('stock_prices.xlsx')
    ws = wb['bitcoin']

    times = ws['A']
    prices = ws['B']

    time_list = []
    price_list = []
    for cell in times:
        time_list.append(cell.value)

    for cell in prices:
        price_list.append(cell.value)

    fig, ax = plt.subplots()

    ax.plot(time_list, price_list)
    ax.set(xlabel = 'Time', ylabel = 'Price', title = 'Bitcoin Price')

    plt.plot(time_list, price_list, '--r')
    plt.show()

update_sheet()
graph_price()